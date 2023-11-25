import tkinter as tk
from tkinter import filedialog, ttk
import csv
import pandas as pd
import os
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

import xlwings as xw


def csv_to_dataframe(file_path):
    return pd.read_csv(file_path)


def read_csv_columns(file_path):
    print(f"Reading columns from: {file_path}")
    with open(file_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        return next(reader, [])  # Return the first row (headers)


def update_common_columns(file1_entry, file2_entry, combobox):
    file1 = file1_entry.get()
    file2 = file2_entry.get()

    if file1 and file2:
        print("Both files selected. Finding common columns...")
        columns_file1 = read_csv_columns(file1)
        columns_file2 = read_csv_columns(file2)

        # Finding common columns
        common_columns = [col for col in columns_file1 if col in columns_file2]
        print(f"Common columns found: {common_columns}")

        # Updating the dropdown menu
        combobox['values'] = common_columns

        if common_columns:
            combobox.current(0)
        else:
            print("No common columns found.")
    else:
        print("Both files are not yet selected.")


def open_file(entry, combobox, other_entry):
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)
        print(f"File selected: {file_path}")
        update_common_columns(file1_entry, file2_entry, combobox)
    else:
        print("File selection cancelled.")


def compare_csvs(file1, file2, key1_index, key2_index):
    print(f"Comparing files: {file1} and {file2}")
    csv1_dict = read_csv_to_dict(file1, key1_index)
    csv2_dict = read_csv_to_dict(file2, key2_index)

    all_keys = set(csv1_dict.keys()) | set(csv2_dict.keys())
    differences = []

    for key in all_keys:
        row1 = csv1_dict.get(key)
        row2 = csv2_dict.get(key)

        if row1 != row2:
            differences.append((key, row1, row2))

    return differences


def read_csv_to_dict(file_path, key_column):
    print(f"Reading file into dictionary: {file_path}")
    data_dict = {}
    with open(file_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            key = row[key_column]
            data_dict[key] = row
    return data_dict


# Modify the report_differences function to call save_to_excel
def report_differences(differences, file1, file2, output_path):
    if not differences:
        print("No differences found.")
    else:
        save_to_excel(file1, file2, differences, output_path)


def compare_and_report():
    file1 = file1_entry.get()
    file2 = file2_entry.get()
    common_key = key_combobox.get()
    output_folder = output_folder_entry.get()

    # Check if both files and a key are selected
    if file1 and file2 and common_key and output_folder:
        print(f"Comparing files: {file1}, {file2} using key: {common_key}")
        output_path = os.path.join(output_folder, "comparison_output.xlsx")
        # Call the save_to_excel function with necessary arguments
        save_to_excel(file1, file2, output_path, common_key)
        print(f"Excel file with comparison saved at: {output_path}")
    else:
        print("Please select both files, a common key, and an output folder")


def save_to_excel(original_file, comparison_file, output_path, common_key):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Original file data
        original_df = csv_to_dataframe(original_file)
        original_df.to_excel(writer, sheet_name="Original file", index=False)

        # Comparison file data
        comparison_df = csv_to_dataframe(comparison_file)
        comparison_df.to_excel(writer, sheet_name="Export file", index=False)

        # Format the 'Compare' sheet as specified
        format_compare_sheet(writer, original_df, comparison_df, common_key)

        print(f"Excel file saved: {output_path}")

    output_folder = output_folder_entry.get()
    output_path = os.path.join(output_folder, "comparison_output.xlsx")
    auto_fill_formula(last_col=get_column_letter(len(original_df.columns) * 3 + 1), last_row=len(original_df)+2, path=output_path)

    # TODO: Needs some formatting - use openpyxl


def format_compare_sheet(writer, original_df, comparison_df, common_key):
    # Create the header with the column names spread across every third column, starting from column B
    header_columns = [''] + sum([[col, '', ''] for col in original_df.columns], [])

    # Initialize the DataFrame with an empty row for the main headers
    compare_df = pd.DataFrame(columns=header_columns)
    compare_df.loc[0] = [''] * len(header_columns)  # Empty row for main headers

    # Prepare the sub-headers "Original", "Export", "Compare" for each column in original_df
    sub_headers = ['Key']
    for _ in original_df.columns:
        sub_headers.extend(['Original', 'Export', 'Compare'])

    # Insert the sub-headers into the second row
    compare_df.loc[0] = sub_headers

    # Add the key data to the DataFrame starting from the third row
    key_data = original_df[common_key].tolist()
    for i, key in enumerate(key_data):
        # Starting index is 2 because we already have headers and sub-headers
        compare_df.loc[i + 1] = [key] + [''] * (len(header_columns) - 1)

    # Writing the 'Compare' DataFrame to the Excel writer
    compare_df.to_excel(writer, sheet_name="Compare", index=False)

    # Access the workbook and the specific sheet to apply formatting
    workbook = writer.book
    worksheet = writer.sheets['Compare']

    # Apply left alignment to the first row
    for cell in worksheet['1:1']:  # Accessing the first row
        cell.alignment = Alignment(horizontal='left')

    set_original_values_formula(worksheet, original_df, comparison_df)



def choose_output_folder(entry):
    folder_path = filedialog.askdirectory()
    if folder_path:
        entry.delete(0, tk.END)
        entry.insert(0, folder_path)
        print(f"Output folder selected: {folder_path}")


def set_original_values_formula(worksheet, original_df, comparison_df):
    key_column_name = key_combobox.get()
    # Calculate the maximum number of rows in both dataframes
    compare_length = len(comparison_df) + 1
    original_length = len(original_df) + 1

    original_key_column = get_column_letter(original_df.columns.get_loc(key_column_name) + 1)
    export_key_column = get_column_letter(comparison_df.columns.get_loc(key_column_name) + 1)

    # Calculate the last row of the DataFrame in the worksheet
    last_row = len(original_df) + 2

    # Define the border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set the formulas for the first row of data (row 3 in the worksheet)
    for idx, col in enumerate(original_df.columns, start=1):
        # Set the HLOOKUP formula for the 'Original' column (every third column starting from B)
        original_col_letter = get_column_letter(idx * 3 - 1)  # Adjust the index for the 'Original' column
        cell = worksheet.cell(row=3, column=idx * 3 - 1)
        cell.value = f"=HLOOKUP(${original_col_letter}$1,'Original file'!$1:${original_length},MATCH($A3,'Original file'!${original_key_column}:${original_key_column},0),FALSE)"

        # Set the HLOOKUP formula for the 'Export' column
        export_col_letter = get_column_letter(idx * 3)  # Adjust the index for the 'Export' column
        cell = worksheet.cell(row=3, column=idx * 3)
        cell.value = f"=HLOOKUP(${original_col_letter}$1,'Export file'!$1:${compare_length},MATCH($A3,'Export file'!${export_key_column}:${export_key_column},0),FALSE)"

        # Set the comparison formula for the 'Compare' column
        compare_col_letter = get_column_letter(idx * 3 + 1)  # Adjust the index for the 'Compare' column
        cell = worksheet.cell(row=3, column=idx * 3 + 1)
        compare_formula = f'=IF(OR(ISNA({original_col_letter}3),ISNA({export_col_letter}3)),"Error",IF({original_col_letter}3={export_col_letter}3,"OK","Error"))'
        cell.value = compare_formula

    # Set the COUNTIF formula in the first row for each 'Compare' column
    for idx in range(4, len(original_df.columns) * 3 + 4, 3):  # Adjust the index for the COUNTIF formula
        col_letter = get_column_letter(idx)
        cell = worksheet.cell(row=1, column=idx)
        cell.value = f'=COUNTIF(${col_letter}$3:${col_letter}{compare_length + 1},"Error")'

    # Apply borders from column B to the end in sections of 3
    for group_start in range(2, len(original_df.columns) * 3 + 2, 3):  # Start from column B (index 2), in steps of 3
        group_end = group_start + 2  # End index of the group

        # Apply borders to each cell in the group
        for row in range(1, last_row + 1):  # Include header and data rows
            for col in range(group_start, group_end + 1):  # From start to end of the group
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                #cell.border = outside_border


def auto_fill_formula(last_col, last_row, path):
    # Open an existing workbook
    wb = xw.Book(path)  # Replace with your file path
    sheet = wb.sheets['Compare']  # Replace with your sheet name

    # Select the cell with the formula
    cell_with_formula = sheet.range(f"B3:{last_col}3")

    # Define the range to auto-fill
    # For example, auto-filling from A2 to A10
    fill_range = sheet.range(f"B3:{last_col}{last_row}")

    # Use the auto-fill
    cell_with_formula.api.AutoFill(fill_range.api, xw.constants.AutoFillType.xlFillDefault)

    # Convert column letter to number
    last_col_index = column_index_from_string(last_col)

    # Apply outside borders to each group of three columns from column B onwards
    for group_start in range(2, last_col_index + 1, 3):  # Start from column B (index 2), in steps of 3
        group_end = group_start + 2  # End index of the group (3 columns per group)

        # Define the range for the current group
        start_col_letter = xw.utils.col_name(group_start)
        end_col_letter = xw.utils.col_name(group_end)

        group_range = sheet.range(f"{start_col_letter}1:{end_col_letter}{last_row}")

        # Apply outside borders to the range using VBA constants
        borders = group_range.api.Borders
        for border_id in [7, 8, 9, 10]:  # These are the VBA constants for xlEdgeLeft, xlEdgeTop, xlEdgeBottom, xlEdgeRight
            border = borders(border_id)
            border.LineStyle = -4119  # xlContinuous
            border.Weight = 2  # xlThin



    # Save and close the workbook
    wb.save()
    # TODO add here the condition if the user wants to close the excel file or not
    wb.close()


# GUI setup
root = tk.Tk()
root.title("CSV Comparison Tool")
# TODO: Add checkbox for leaving the Excel window open
# TODO: Open Folder button for the output folder

# File selection for File 1
file1_frame = tk.Frame(root)
file1_frame.pack(padx=10, pady=5)
file1_entry = tk.Entry(file1_frame, width=50)
file1_entry.pack(side=tk.LEFT)
file1_button = tk.Button(file1_frame, text="Select File 1", command=lambda: open_file(file1_entry, key_combobox, file2_entry))
file1_button.pack(side=tk.LEFT)

# File selection for File 2
file2_frame = tk.Frame(root)
file2_frame.pack(padx=10, pady=5)
file2_entry = tk.Entry(file2_frame, width=50)
file2_entry.pack(side=tk.LEFT)
file2_button = tk.Button(file2_frame, text="Select File 2", command=lambda: open_file(file2_entry, key_combobox, file1_entry))
file2_button.pack(side=tk.LEFT)

# Key selection (common columns)
key_frame = tk.Frame(root)
key_frame.pack(padx=10, pady=5)
key_combobox = ttk.Combobox(key_frame, width=15)
key_combobox.pack(side=tk.LEFT)

# Output folder selection
output_folder_frame = tk.Frame(root)
output_folder_frame.pack(padx=10, pady=5)
output_folder_entry = tk.Entry(output_folder_frame, width=50)
output_folder_entry.pack(side=tk.LEFT)
output_folder_button = tk.Button(output_folder_frame, text="Select Output Folder", command=lambda: choose_output_folder(output_folder_entry))
output_folder_button.pack(side=tk.LEFT)

# 'Compare Files' button
compare_button = tk.Button(root, text="Compare Files", command=compare_and_report)
compare_button.pack(pady=10)

# Start the GUI
root.mainloop()
